import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image
import os
import openpyxl
from openpyxl.utils import get_column_letter

def select_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        image_folder_path.set(folder_path)

def extract_brightness():
    try:
        # X座標、Y座標、スケールの入力処理
        x_coords = x_entry.get().split(',')
        y_origins = y_entry.get().split(',')
        scales = scale_entry.get().split(',')

        if not (len(x_coords) == len(y_origins) == len(scales)):
            messagebox.showerror("Error", "X座標、Y座標、スケールの数が一致していません。")
            return

        x_coords = [int(x.strip()) for x in x_coords]
        y_origins = [int(y.strip()) for y in y_origins]
        scales = [float(s.strip()) for s in scales]
        folder_path_val = image_folder_path.get()

        if not os.path.exists(folder_path_val):
            messagebox.showerror("Error", "選択されたフォルダが存在しません。")
            return

        image_files = [f for f in os.listdir(folder_path_val) if f.lower().endswith((".png", ".jpg", ".jpeg", ".bmp", ".tiff"))]
        if not image_files:
            messagebox.showerror("Error", "選択されたフォルダに画像ファイルが見つかりません。")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if not save_path:
            return

        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # デフォルトのシートを削除

        for image_file in image_files:
            image_path = os.path.join(folder_path_val, image_file)
            img = Image.open(image_path).convert("L")  # グレースケールに変換
            width, height = img.size

            # 有効なX座標、Y原点、スケールをフィルタリング
            valid_coords_data = [(x, y_origin, scale) for x, y_origin, scale in zip(x_coords, y_origins, scales) if 0 <= x < width and 0 <= y_origin < height]
            if not valid_coords_data:
                messagebox.showwarning("Warning", f"{image_file}をスキップ: 有効な座標が画像の範囲外です。")
                continue

            sheet = workbook.create_sheet(title=os.path.splitext(image_file)[0])

            # ラベルを「輝度 → スケール」の順に並べる
            header = ["y"]
            for x, _, _ in valid_coords_data:
                header.append(f"Brightness x={x}")
                header.append(f"Scale (mm) x={x}")
            sheet.append(header)

            # データを「輝度 → スケール」の順に並べる
            for y in range(height):
                row = [y]
                for x, y_origin, scale in valid_coords_data:
                    brightness = img.getpixel((x, y))
                    physical_length = (y_origin - y) * scale  # Y原点からの距離（上が正、下が負）
                    row.extend([brightness, physical_length])
                sheet.append(row)

            # 列幅の自動調整
            for col_num, _ in enumerate(sheet.iter_cols(), 1):
                column_width = max(len(str(cell.value)) for cell in sheet[get_column_letter(col_num)])
                sheet.column_dimensions[get_column_letter(col_num)].width = column_width + 2

        workbook.save(save_path)
        messagebox.showinfo("Success", f"輝度値とスケールが{save_path}に保存されました")

    except ValueError:
        messagebox.showerror("Error", "X座標、Y座標、およびスケールには有効な数値を入力してください。")
    except Exception as e:
        messagebox.showerror("Error", f"予期しないエラーが発生しました: {e}")

# GUI Setup
root = tk.Tk()
root.title("Batch Brightness Extractor with Scale and Origin")

image_folder_path = tk.StringVar()

frame = tk.Frame(root, padx=10, pady=10)
frame.pack(fill="both", expand=True)

# フォルダ選択
tk.Label(frame, text="Select Folder:").grid(row=0, column=0, sticky="e", pady=5)
tk.Entry(frame, textvariable=image_folder_path, width=40).grid(row=0, column=1, padx=5, pady=5)
tk.Button(frame, text="Browse", command=select_folder).grid(row=0, column=2, pady=5)

# X座標入力
tk.Label(frame, text="X-Coordinates (comma-separated):").grid(row=1, column=0, sticky="e", pady=5)
x_entry = tk.Entry(frame, width=30)
x_entry.grid(row=1, column=1, padx=5, pady=5)

# Y座標入力
tk.Label(frame, text="Y-Coordinates (comma-separated):").grid(row=2, column=0, sticky="e", pady=5)
y_entry = tk.Entry(frame, width=30)
y_entry.grid(row=2, column=1, padx=5, pady=5)

# スケール入力
tk.Label(frame, text="Scale (mm/pixel, comma-separated):").grid(row=3, column=0, sticky="e", pady=5)
scale_entry = tk.Entry(frame, width=30)
scale_entry.grid(row=3, column=1, padx=5, pady=5)

# 抽出と保存ボタン
tk.Button(frame, text="Extract Brightness", command=extract_brightness).grid(row=4, column=0, columnspan=3, pady=10)

root.mainloop()
