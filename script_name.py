import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import re
from datetime import datetime
import csv

def select_file():
    """ファイル選択ダイアログを開く"""
    file_path = filedialog.askopenfilename(
        title="Excelファイルを選択",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    entry_file_path.delete(0, tk.END)
    entry_file_path.insert(0, file_path)

def extract_values():
    """複数セルの値を時系列順に並べて抽出"""
    file_path = entry_file_path.get()
    cells = entry_cells.get().split(",")
    
    if not file_path:
        messagebox.showerror("エラー", "ファイルを選択してください")
        return
    
    if not cells:
        messagebox.showerror("エラー", "セル番号を入力してください")
        return
    
    try:
        # Excelファイルを開く（data_only=Trueで数式の結果を取得）
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # シート名を時系列順に並べ替える
        sheet_names = sorted(
            workbook.sheetnames,
            key=lambda name: datetime.strptime(re.match(r"(\d{14})", name).group(1), "%Y%m%d%H%M%S")
        )
        
        # 結果を保存するリスト
        result = [["シート名"] + cells]
        
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            row = [sheet_name]
            for cell in cells:
                value = sheet[cell.strip()].value
                row.append(value if value is not None else "")
            result.append(row)
        
        # 結果をGUIに表示
        display_result(result)
        
        # CSVエクスポートボタンとExcelエクスポートボタンを有効化
        export_button_csv.config(state=tk.NORMAL)
        export_button_excel.config(state=tk.NORMAL)
        global extracted_result
        extracted_result = result  # エクスポート用に結果を保存
    
    except FileNotFoundError:
        messagebox.showerror("エラー", "指定したファイルが見つかりません")
    except KeyError:
        messagebox.showerror("エラー", "無効なセル番号です")
    except Exception as e:
        messagebox.showerror("エラー", f"予期しないエラー: {e}")

def display_result(result):
    """結果をテキストウィジェットに表示"""
    result_text.delete(1.0, tk.END)
    
    for row in result:
        result_text.insert(tk.END, "\t".join(map(str, row)) + "\n")

def export_to_csv():
    """抽出結果をCSVファイルにエクスポート"""
    if not extracted_result:
        messagebox.showerror("エラー", "抽出された結果がありません")
        return
    
    file_path = filedialog.asksaveasfilename(
        title="CSVファイルを保存",
        defaultextension=".csv",
        filetypes=[("CSV files", "*.csv")]
    )
    
    if not file_path:
        return  # キャンセル時は何もしない
    
    try:
        with open(file_path, mode="w", newline="", encoding="utf-8-sig") as file:
            writer = csv.writer(file)
            writer.writerows(extracted_result)
        messagebox.showinfo("完了", "CSVファイルを正常に保存しました")
    except Exception as e:
        messagebox.showerror("エラー", f"CSVファイルの保存中にエラーが発生しました: {e}")

def export_to_excel():
    """抽出結果をExcelファイルにエクスポート"""
    if not extracted_result:
        messagebox.showerror("エラー", "抽出された結果がありません")
        return
    
    file_path = filedialog.asksaveasfilename(
        title="Excelファイルを保存",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    
    if not file_path:
        return  # キャンセル時は何もしない
    
    try:
        # 新しいExcelワークブックを作成
        new_workbook = openpyxl.Workbook()
        sheet = new_workbook.active
        sheet.title = "抽出結果"
        
        # 抽出結果を書き込む
        for row in extracted_result:
            sheet.append(row)
        
        # Excelファイルを保存
        new_workbook.save(file_path)
        messagebox.showinfo("完了", "Excelファイルを正常に保存しました")
    except Exception as e:
        messagebox.showerror("エラー", f"Excelファイルの保存中にエラーが発生しました: {e}")

# GUI構築
root = tk.Tk()
root.title("Excelセル抽出ツール（時系列対応）")

# ファイルパス入力
frame_file = tk.Frame(root)
frame_file.pack(pady=10)

tk.Label(frame_file, text="Excelファイル:").pack(side=tk.LEFT)
entry_file_path = tk.Entry(frame_file, width=50)
entry_file_path.pack(side=tk.LEFT, padx=5)
tk.Button(frame_file, text="参照", command=select_file).pack(side=tk.LEFT)

# セル番号入力
frame_cells = tk.Frame(root)
frame_cells.pack(pady=10)

tk.Label(frame_cells, text="セル番号 (カンマ区切り, 例: A1, B2, C3):").pack(side=tk.LEFT)
entry_cells = tk.Entry(frame_cells, width=30)
entry_cells.pack(side=tk.LEFT, padx=5)

# 抽出ボタン
extract_button = tk.Button(root, text="抽出", command=extract_values)
extract_button.pack(pady=10)

# CSVエクスポートボタン
export_button_csv = tk.Button(root, text="CSVにエクスポート", command=export_to_csv, state=tk.DISABLED)
export_button_csv.pack(pady=5)

# Excelエクスポートボタン
export_button_excel = tk.Button(root, text="Excelにエクスポート", command=export_to_excel, state=tk.DISABLED)
export_button_excel.pack(pady=5)

# 結果表示テキスト
result_text = tk.Text(root, width=80, height=20)
result_text.pack(pady=10)

root.mainloop()
