# file path: gui_brightness_integrator.py
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from concurrent.futures import ProcessPoolExecutor

def process_excel():
    root = tk.Tk()
    root.title("Brightness Integrator")

    selected_file_path = tk.StringVar(value="No file selected")

    def select_file():
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if file_path:
            selected_file_path.set(file_path)

    def calculate_brightness(file_path, sheet_name):
        """シートごとに輝度積分を計算する関数（マルチプロセス用）"""
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook[sheet_name]
        total_brightness_area = 0

        if sheet.max_row < 2 or sheet.max_column < 2:
            return sheet_name, total_brightness_area

        for row in sheet.iter_rows(min_row=2, min_col=2, values_only=True):
            for brightness in row:
                if brightness is not None:
                    total_brightness_area += brightness  # ピクセルごとの輝度値を加算
        return sheet_name, total_brightness_area

    def calculate_and_save():
        try:
            file_path = selected_file_path.get()
            if file_path == "No file selected":
                messagebox.showerror("Error", "Please select an Excel file first!")
                return

            # Excelファイルのシート名を取得
            workbook = openpyxl.load_workbook(file_path)
            sheet_names = workbook.sheetnames

            # マルチプロセスで各シートの計算を実行
            with ProcessPoolExecutor() as executor:
                futures = [
                    executor.submit(calculate_brightness, file_path, sheet_name)
                    for sheet_name in sheet_names
                ]
                results = [future.result() for future in futures]

            # 結果を新しいワークブックに保存
            output_workbook = openpyxl.Workbook()
            output_sheet = output_workbook.active
            output_sheet.title = "Integrated Results"
            output_sheet.append(["Sheet Name", "Integrated Brightness (brightness × pixels²)"])

            for name, area in results:
                output_sheet.append([name, area])

            # 保存ダイアログを表示
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
            )
            if not save_path:
                return
            output_workbook.save(save_path)
            messagebox.showinfo("Success", f"Results saved to {save_path}")
            root.destroy()

        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")

    # GUI要素の作成
    tk.Label(root, text="Select an Excel file to integrate brightness").grid(row=0, column=0, columnspan=2, pady=10)

    # ファイル選択ボタンとラベル
    tk.Button(root, text="Select Excel File", command=select_file).grid(row=1, column=0, padx=5, pady=5)
    tk.Label(root, textvariable=selected_file_path, wraplength=300).grid(row=1, column=1, padx=5, pady=5)

    # 計算実行ボタン
    calculate_button = ttk.Button(root, text="Calculate and Save", command=calculate_and_save)
    calculate_button.grid(row=2, column=0, columnspan=2, pady=10)

    root.mainloop()

if __name__ == "__main__":
    process_excel()
