# File path: batch_brightness_extractor_rectangle_with_scale.py
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image
import os
import openpyxl
from openpyxl.utils import get_column_letter

def select_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        image_folder_path.set(folder_path)

def extract_brightness_rectangle():
    try:
        # 座標と矩形サイズ、スケールの入力処理
        x_coord = int(x_entry.get().strip())
        y_coord = int(y_entry.get().strip())
        rect_width = int(width_entry.get().strip())
        rect_height = int(height_entry.get().strip())
        scale_x = float(scale_x_entry.get().strip())
        scale_y = float(scale_y_entry.get().strip())
        folder_path_val = image_folder_path.get()

        if not os.path.exists(folder_path_val):
            messagebox.showerror("Error", "選択されたフォルダが存在しません。")
            return

        image_files = [f for f in os.listdir(folder_path_val) if f.lower().endswith((".png", ".jpg", ".jpeg", ".bmp", ".tiff"))]
        if not image_files:
            messagebox.showerror("Error", "選択されたフォルダに画像ファイルが見つかりません。")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if not save_path:
            return

        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # デフォルトのシートを削除

        for image_file in image_files:
            image_path = os.path.join(folder_path_val, image_file)
            img = Image.open(image_path).convert("L")  # グレースケールに変換
            width, height = img.size

            # 入力範囲が画像サイズを超えていないかチェック
            if not (0 <= x_coord < width and 0 <= y_coord < height):
                messagebox.showwarning("Warning", f"{image_file}をスキップ: 指定された座標が画像範囲外です。")
                continue
            if x_coord + rect_width > width or y_coord + rect_height > height:
                messagebox.showwarning("Warning", f"{image_file}をスキップ: 矩形範囲が画像の範囲外です。")
                continue

            # 輝度値を切り出してExcelに保存
            sheet = workbook.create_sheet(title=os.path.splitext(image_file)[0])
            header = ["y\\x", "Physical y (mm)"]
            for i in range(rect_width):
                header.extend([f"Brightness x={x_coord + i}", f"Physical x (mm) x={x_coord + i}"])
            sheet.append(header)

            for j in range(rect_height):
                row = [y_coord + j, j * scale_y]  # Y物理座標
                for i in range(rect_width):
                    brightness = img.getpixel((x_coord + i, y_coord + j))
                    physical_x = i * scale_x  # X物理座標
                    row.extend([brightness, physical_x])
                sheet.append(row)

            # 列幅の自動調整
            for col_num, _ in enumerate(sheet.iter_cols(), 1):
                column_width = max(len(str(cell.value)) for cell in sheet[get_column_letter(col_num)])
                sheet.column_dimensions[get_column_letter(col_num)].width = column_width + 2

        workbook.save(save_path)
        messagebox.showinfo("Success", f"輝度値が{save_path}に保存されました")

    except ValueError:
        messagebox.showerror("Error", "座標、矩形のサイズ、およびスケールには有効な数値を入力してください。")
    except Exception as e:
        messagebox.showerror("Error", f"予期しないエラーが発生しました: {e}")

# GUI Setup
root = tk.Tk()
root.title("Rectangle Brightness Extractor with X and Y Scales")

image_folder_path = tk.StringVar()

frame = tk.Frame(root, padx=10, pady=10)
frame.pack(fill="both", expand=True)

# フォルダ選択
tk.Label(frame, text="Select Folder:").grid(row=0, column=0, sticky="e", pady=5)
tk.Entry(frame, textvariable=image_folder_path, width=40).grid(row=0, column=1, padx=5, pady=5)
tk.Button(frame, text="Browse", command=select_folder).grid(row=0, column=2, pady=5)

# X座標入力
tk.Label(frame, text="X-Coordinate:").grid(row=1, column=0, sticky="e", pady=5)
x_entry = tk.Entry(frame, width=30)
x_entry.grid(row=1, column=1, padx=5, pady=5)

# Y座標入力
tk.Label(frame, text="Y-Coordinate:").grid(row=2, column=0, sticky="e", pady=5)
y_entry = tk.Entry(frame, width=30)
y_entry.grid(row=2, column=1, padx=5, pady=5)

# 幅入力
tk.Label(frame, text="Rectangle Width:").grid(row=3, column=0, sticky="e", pady=5)
width_entry = tk.Entry(frame, width=30)
width_entry.grid(row=3, column=1, padx=5, pady=5)

# 高さ入力
tk.Label(frame, text="Rectangle Height:").grid(row=4, column=0, sticky="e", pady=5)
height_entry = tk.Entry(frame, width=30)
height_entry.grid(row=4, column=1, padx=5, pady=5)

# スケール入力 (X方向)
tk.Label(frame, text="Scale X (mm/pixel):").grid(row=5, column=0, sticky="e", pady=5)
scale_x_entry = tk.Entry(frame, width=30)
scale_x_entry.grid(row=5, column=1, padx=5, pady=5)

# スケール入力 (Y方向)
tk.Label(frame, text="Scale Y (mm/pixel):").grid(row=6, column=0, sticky="e", pady=5)
scale_y_entry = tk.Entry(frame, width=30)
scale_y_entry.grid(row=6, column=1, padx=5, pady=5)

# 抽出と保存ボタン
tk.Button(frame, text="Extract Brightness", command=extract_brightness_rectangle).grid(row=7, column=0, columnspan=3, pady=10)

root.mainloop()
